"""엑셀 내보내기 노드 — 시스템 데이터를 표준 양식 .xlsx 파일로 저장"""
import logging
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ..registry import NodeHandlerRegistry
from ..base import NodeHandler, ExecutionContext

logger = logging.getLogger(__name__)

# 기본 export 디렉토리: backend/data/exports
_THIS_FILE = Path(__file__).resolve()
# excel_export.py → action → nodes → app → backend
_BACKEND_DIR = _THIS_FILE.parents[3]
_DEFAULT_EXPORT_DIR = _BACKEND_DIR / "data" / "exports"


def _safe_filename(name: str) -> str:
    name = re.sub(r"[\\/:*?\"<>|]", "_", name or "").strip()
    return name or "export"


def _coerce_rows(input_data: Any) -> List[Dict[str, Any]]:
    """입력에서 rows 리스트를 추출한다."""
    if input_data is None:
        return []

    # dict 입력
    if isinstance(input_data, dict):
        # 명시적 rows 키 우선
        if "rows" in input_data and isinstance(input_data["rows"], list):
            return [r for r in input_data["rows"] if isinstance(r, dict)]
        # data 키 (api-call 결과가 {status, data} 형태일 수 있음)
        if "data" in input_data:
            return _coerce_rows(input_data["data"])
        # dict 자체가 단일 row
        return [input_data]

    # list 입력
    if isinstance(input_data, list):
        return [r for r in input_data if isinstance(r, dict)]

    return []


def _resolve_value(row: Dict[str, Any], key: str) -> Any:
    """dot-path로 중첩 값 조회. 없으면 빈 문자열."""
    if not key:
        return ""
    parts = key.split(".")
    cur: Any = row
    for p in parts:
        if isinstance(cur, dict) and p in cur:
            cur = cur[p]
        else:
            return ""
    if cur is None:
        return ""
    if isinstance(cur, (dict, list)):
        import json
        try:
            return json.dumps(cur, ensure_ascii=False)
        except Exception:
            return str(cur)
    return cur


@NodeHandlerRegistry.register
class ExcelExportHandler(NodeHandler):
    node_type = "excel-export"
    category = "action"
    display_name = "엑셀 내보내기"
    description = "시스템 데이터를 표준 엑셀(.xlsx) 파일로 저장합니다"

    async def execute(
        self,
        node: Any,
        input_data: Dict[str, Any],
        ctx: ExecutionContext,
    ) -> Any:
        config = node.config or {}
        sheet_name: str = (config.get("sheetName") or "Sheet1").strip() or "Sheet1"
        columns_cfg: List[Dict[str, str]] = config.get("columns") or []
        template_doc_id: Optional[str] = config.get("templateDocId") or None
        output_path_cfg: Optional[str] = config.get("outputPath") or None

        # 입력에서 rows / filename 추출
        filename_hint = "export"
        if isinstance(input_data, dict):
            filename_hint = str(input_data.get("filename") or filename_hint)

        rows = _coerce_rows(input_data)

        # columns 자동 추론: 설정이 비어 있으면 첫 row의 키 사용
        if not columns_cfg:
            first = rows[0] if rows else {}
            columns_cfg = [{"header": k, "key": k} for k in first.keys()]

        # 출력 경로 결정
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if output_path_cfg:
            out = ctx.render_template(output_path_cfg, input_data) if ctx and ctx.render_template else output_path_cfg
            output_path = Path(out)
        else:
            safe_name = _safe_filename(filename_hint)
            output_path = _DEFAULT_EXPORT_DIR / f"{timestamp}_{safe_name}.xlsx"

        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Workbook 생성
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # 엑셀 시트명 31자 제한

        # 헤더 작성
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        headers = [c.get("header", c.get("key", "")) for c in columns_cfg]
        keys = [c.get("key", c.get("header", "")) for c in columns_cfg]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # 데이터 작성
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, key in enumerate(keys, start=1):
                val = _resolve_value(row, key)
                # datetime 등 처리
                if isinstance(val, datetime):
                    val = val.strftime("%Y-%m-%d %H:%M:%S")
                ws.cell(row=row_idx, column=col_idx, value=val)

        # 자동 컬럼 너비
        for col_idx, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for row in rows:
                v = _resolve_value(row, keys[col_idx - 1])
                vl = len(str(v)) if v is not None else 0
                if vl > max_len:
                    max_len = vl
            width = min(max(int(max_len * 1.2) + 2, 10), 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        wb.save(str(output_path))
        row_count = len(rows)
        logger.info(f"Excel exported: {output_path} ({row_count} rows)")

        try:
            rel = output_path.name
            url = f"/api/v1/files/exports/{rel}"
        except Exception:
            url = None

        return {
            "file_path": str(output_path),
            "row_count": row_count,
            "sheet_name": ws.title,
            "columns": headers,
            "template_doc_id": template_doc_id,
            "url": url,
        }
