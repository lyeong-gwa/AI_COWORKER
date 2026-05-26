"""record_exporter — InstanceDB record 즉석 변환 모듈.

포맷별로 record dict 를 bytes 로 변환하여 (content, mime_type, filename) 튜플을 반환.
디스크 저장 없이 모두 메모리 스트림으로 처리한다.

지원 포맷:
  md   — Markdown (GFM 원문 그대로 / 정의 목록)
  csv  — UTF-8 BOM, Excel 한글 호환
  html — 완전한 HTML 문서, markdown 필드 GFM 렌더링
  xlsx — openpyxl 의존. 없으면 ImportError 발생 (라우트가 503 처리)
"""

from __future__ import annotations

import csv
import html as html_lib
import io
from typing import Dict, Optional, Any


# ── 마크다운 렌더러 ─────────────────────────────────────────────────────────

def _render_markdown(text: str) -> str:
    """GFM 텍스트를 HTML 로 변환. markdown 패키지 우선, 없으면 escape 후 <pre>."""
    try:
        import markdown as md_lib
        return md_lib.markdown(text, extensions=["tables", "fenced_code"])
    except ImportError:
        pass
    return f"<pre>{html_lib.escape(text)}</pre>"


# ── 값 → 문자열 ─────────────────────────────────────────────────────────────

def _to_str(value: Any) -> str:
    if isinstance(value, str):
        return value
    import json
    return json.dumps(value, ensure_ascii=False)


# ── md 변환 ─────────────────────────────────────────────────────────────────

def _export_md(
    record: dict,
    field: Optional[str],
    viewer_hints: Optional[Dict[str, str]],
) -> bytes:
    hints = viewer_hints or {}
    data: dict = record.get("data") or {}

    if field is not None:
        value = data.get(field, "")
        hint = hints.get(field, "")
        raw = _to_str(value)
        if hint == "markdown":
            content = raw
        else:
            content = f"## {field}\n\n{raw}"
        return content.encode("utf-8")

    # 전체 record 마크다운 직렬화
    lines: list[str] = []
    rec_id = record.get("id", "unknown")
    lines.append(f"# Record `{rec_id}`\n")

    src = record.get("_source") or {}
    if any(src.values()):
        lines.append("**출처**")
        for k, v in src.items():
            if v:
                lines.append(f"- {k}: `{v}`")
        lines.append("")

    created_at = record.get("createdAt", "")
    if created_at:
        lines.append(f"**생성일시**: {created_at}\n")

    lines.append("---\n")

    for key, value in data.items():
        hint = hints.get(key, "")
        lines.append(f"### {key}\n")
        if hint == "markdown":
            lines.append(_to_str(value))
        else:
            lines.append(f"- **{key}**: {_to_str(value)}")
        lines.append("")

    return "\n".join(lines).encode("utf-8")


# ── csv 변환 ─────────────────────────────────────────────────────────────────

def _export_csv(
    record: dict,
    field: Optional[str],
    viewer_hints: Optional[Dict[str, str]],
) -> bytes:
    data: dict = record.get("data") or {}
    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_ALL)

    if field is not None:
        value = data.get(field, "")
        writer.writerow([field])
        writer.writerow([_to_str(value)])
    else:
        keys = list(data.keys())
        writer.writerow(keys)
        writer.writerow([_to_str(data[k]) for k in keys])

    content = buf.getvalue()
    # UTF-8 BOM for Excel 한글 호환
    return b"\xef\xbb\xbf" + content.encode("utf-8")


# ── html 변환 ────────────────────────────────────────────────────────────────

_HTML_STYLE = """
<style>
  body { font-family: -apple-system, sans-serif; max-width: 900px; margin: 2rem auto; padding: 0 1rem; color: #222; }
  h1 { font-size: 1.4rem; border-bottom: 2px solid #e0e0e0; padding-bottom: .4rem; }
  h2 { font-size: 1.1rem; color: #444; margin-top: 1.5rem; }
  dl { margin: 0; }
  dt { font-weight: 600; font-size: .85rem; color: #666; margin-top: 1rem; text-transform: uppercase; letter-spacing: .05em; }
  dd { margin-left: 0; padding: .5rem; background: #f7f7f7; border-radius: 4px; word-break: break-all; }
  table { border-collapse: collapse; width: 100%; margin: .5rem 0; }
  th, td { border: 1px solid #d0d0d0; padding: .4rem .7rem; text-align: left; font-size: .9rem; }
  th { background: #f0f0f0; }
  pre, code { background: #f4f4f4; border-radius: 3px; padding: .15rem .35rem; font-size: .85rem; }
  pre { padding: .7rem 1rem; overflow-x: auto; }
  .meta { font-size: .8rem; color: #999; margin-bottom: 1rem; }
</style>
"""


def _export_html(
    record: dict,
    field: Optional[str],
    viewer_hints: Optional[Dict[str, str]],
) -> bytes:
    hints = viewer_hints or {}
    data: dict = record.get("data") or {}
    rec_id = record.get("id", "unknown")

    if field is not None:
        value = data.get(field, "")
        hint = hints.get(field, "")
        raw = _to_str(value)
        if hint == "markdown":
            body_html = f"<h2>{html_lib.escape(field)}</h2>" + _render_markdown(raw)
        else:
            body_html = f"<h2>{html_lib.escape(field)}</h2><dd>{html_lib.escape(raw)}</dd>"
    else:
        parts: list[str] = ["<dl>"]
        for key, value in data.items():
            hint = hints.get(key, "")
            raw = _to_str(value)
            parts.append(f"<dt>{html_lib.escape(key)}</dt>")
            if hint == "markdown":
                parts.append(f"<dd>{_render_markdown(raw)}</dd>")
            else:
                parts.append(f"<dd>{html_lib.escape(raw)}</dd>")
        parts.append("</dl>")
        body_html = "\n".join(parts)

    created_at = record.get("createdAt", "")
    meta_line = f'<p class="meta">id: {html_lib.escape(rec_id)}'
    if created_at:
        meta_line += f" &nbsp;·&nbsp; {html_lib.escape(created_at)}"
    meta_line += "</p>"

    html_doc = (
        f"<!DOCTYPE html>\n"
        f"<html lang='ko'>\n"
        f"<head>\n"
        f"<meta charset='utf-8'>\n"
        f"<meta name='viewport' content='width=device-width,initial-scale=1'>\n"
        f"<title>Record {html_lib.escape(rec_id)}</title>\n"
        f"{_HTML_STYLE}"
        f"</head>\n"
        f"<body>\n"
        f"<h1>Record <code>{html_lib.escape(rec_id)}</code></h1>\n"
        f"{meta_line}\n"
        f"{body_html}\n"
        f"</body>\n"
        f"</html>\n"
    )
    return html_doc.encode("utf-8")


# ── xlsx 변환 ────────────────────────────────────────────────────────────────

def _export_xlsx(
    record: dict,
    field: Optional[str],
    viewer_hints: Optional[Dict[str, str]],
) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as e:
        raise ImportError("openpyxl 패키지가 설치되어 있지 않습니다") from e

    data: dict = record.get("data") or {}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Record"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D0E8FF", end_color="D0E8FF", fill_type="solid")

    if field is not None:
        value = data.get(field, "")
        # 헤더
        c_header = ws.cell(row=1, column=1, value=field)
        c_header.font = header_font
        c_header.fill = header_fill
        # 값
        ws.cell(row=2, column=1, value=_to_str(value))
        ws.column_dimensions["A"].width = max(len(field) + 4, 20)
    else:
        keys = list(data.keys())
        for col_idx, key in enumerate(keys, start=1):
            c = ws.cell(row=1, column=col_idx, value=key)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(wrap_text=False)

            val_str = _to_str(data[key])
            ws.cell(row=2, column=col_idx, value=val_str)

            # 한글 포함 칼럼 width 자동 조정 (대략적)
            col_width = max(
                _char_width(key),
                _char_width(val_str[:50]),  # 최대 50자 기준
            ) + 4
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = min(col_width, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _char_width(text: str) -> int:
    """한글은 2칸, 그 외 1칸으로 계산한 대략적 칼럼 너비."""
    width = 0
    for ch in text:
        if "가" <= ch <= "힣" or "一" <= ch <= "鿿":
            width += 2
        else:
            width += 1
    return width


# ── 공개 API ─────────────────────────────────────────────────────────────────

_SUPPORTED_FORMATS = {"md", "csv", "html", "xlsx"}

_MIME = {
    "md": "text/markdown; charset=utf-8",
    "csv": "text/csv; charset=utf-8",
    "html": "text/html; charset=utf-8",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}

_EXT = {
    "md": "md",
    "csv": "csv",
    "html": "html",
    "xlsx": "xlsx",
}


def export_record(
    record: dict,
    fmt: str,
    *,
    field: Optional[str] = None,
    viewer_hints: Optional[Dict[str, str]] = None,
) -> tuple[bytes, str, str]:
    """record(JSON dict) → (content_bytes, mime_type, suggested_filename).

    Parameters
    ----------
    record:
        instance_db_store 에서 읽어온 raw record dict.
        {"id": ..., "data": {...}, "_source": {...}, "createdAt": ...}
    fmt:
        "md" | "csv" | "html" | "xlsx"
    field:
        record.data 의 특정 키. 지정 시 해당 필드만 변환.
    viewer_hints:
        컬렉션 메타의 viewerHints dict (필드명 → 힌트 타입).
        md/html 에서 "markdown" 힌트 필드를 GFM 렌더링에 활용.
    """
    if fmt not in _SUPPORTED_FORMATS:
        raise ValueError(f"지원하지 않는 포맷: {fmt!r}. 지원 포맷: {sorted(_SUPPORTED_FORMATS)}")

    rec_id = record.get("id") or "record"
    filename = f"{rec_id}.{_EXT[fmt]}"

    if fmt == "md":
        content = _export_md(record, field, viewer_hints)
    elif fmt == "csv":
        content = _export_csv(record, field, viewer_hints)
    elif fmt == "html":
        content = _export_html(record, field, viewer_hints)
    else:  # xlsx
        content = _export_xlsx(record, field, viewer_hints)

    return content, _MIME[fmt], filename


__all__ = ["export_record"]
