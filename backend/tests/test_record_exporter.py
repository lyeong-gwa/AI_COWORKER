"""record_exporter 유닛테스트.

검증 포인트:
1. md 변환 — field 지정 / 미지정 (markdown hint 포함)
2. csv 변환 — UTF-8 BOM 검증
3. html 변환 — markdown 힌트 필드가 GFM 태그로 렌더링되는지
4. xlsx 변환 — openpyxl 로 다시 열어 셀 값 확인
"""

import io
import pytest

from app.services.record_exporter import export_record

# ── 공통 픽스처 ─────────────────────────────────────────────────────────────

SAMPLE_RECORD = {
    "id": "rec-abcd1234",
    "data": {
        "title": "샘플 제목",
        "answer_md": "# 안녕하세요\n\n**굵은 텍스트**와 `코드`입니다.",
        "score": 42,
    },
    "_source": {
        "workflowId": "wf-001",
        "executionId": "ex-001",
        "warehouseId": None,
    },
    "createdAt": "2026-05-12T10:00:00",
}

VIEWER_HINTS = {
    "answer_md": "markdown",
    "title": "text",
}


# ── md 변환 ─────────────────────────────────────────────────────────────────

class TestMdExport:
    def test_field_markdown_hint_returns_raw(self):
        """markdown hint 필드는 GFM 원문 그대로 반환."""
        content, mime, filename = export_record(
            SAMPLE_RECORD, "md", field="answer_md", viewer_hints=VIEWER_HINTS
        )
        text = content.decode("utf-8")
        # 원문 그대로 (헤더/마크업 유지)
        assert "# 안녕하세요" in text
        assert "**굵은 텍스트**" in text
        assert mime == "text/markdown; charset=utf-8"
        assert filename.endswith(".md")

    def test_field_non_markdown_wraps_in_header(self):
        """markdown hint 아닌 필드는 ## 헤더 + 값으로 반환."""
        content, _, _ = export_record(
            SAMPLE_RECORD, "md", field="title", viewer_hints=VIEWER_HINTS
        )
        text = content.decode("utf-8")
        assert "## title" in text
        assert "샘플 제목" in text

    def test_no_field_serializes_all_keys(self):
        """field 미지정 시 모든 data 키가 포함되어야 한다."""
        content, mime, _ = export_record(
            SAMPLE_RECORD, "md", viewer_hints=VIEWER_HINTS
        )
        text = content.decode("utf-8")
        assert "title" in text
        assert "answer_md" in text
        assert "score" in text
        # markdown hint 필드의 GFM 원문이 inline
        assert "# 안녕하세요" in text
        assert mime == "text/markdown; charset=utf-8"

    def test_record_id_in_output(self):
        """전체 직렬화 시 record id가 포함된다."""
        content, _, _ = export_record(SAMPLE_RECORD, "md")
        assert "rec-abcd1234" in content.decode("utf-8")


# ── csv 변환 ─────────────────────────────────────────────────────────────────

class TestCsvExport:
    def test_utf8_bom_present(self):
        """UTF-8 BOM (0xEF BB BF) 이 앞에 붙어 있어야 한다."""
        content, _, _ = export_record(SAMPLE_RECORD, "csv")
        assert content[:3] == b"\xef\xbb\xbf", "UTF-8 BOM 없음"

    def test_headers_and_values_present(self):
        """첫 행 = 키, 두 번째 행 = 값."""
        content, mime, filename = export_record(SAMPLE_RECORD, "csv")
        text = content.decode("utf-8-sig")  # BOM 제거
        lines = [l for l in text.splitlines() if l.strip()]
        assert len(lines) >= 2
        header_line = lines[0]
        assert "title" in header_line
        assert "answer_md" in header_line
        assert mime == "text/csv; charset=utf-8"
        assert filename.endswith(".csv")

    def test_field_single_column(self):
        """field 지정 시 1컬럼 1행."""
        content, _, _ = export_record(SAMPLE_RECORD, "csv", field="title")
        text = content.decode("utf-8-sig")
        lines = [l for l in text.splitlines() if l.strip()]
        assert len(lines) == 2
        assert "title" in lines[0]
        assert "샘플 제목" in lines[1]

    def test_korean_value_round_trip(self):
        """한글 값이 BOM decode 후 원문과 일치."""
        content, _, _ = export_record(SAMPLE_RECORD, "csv", field="title")
        text = content.decode("utf-8-sig")
        assert "샘플 제목" in text


# ── html 변환 ────────────────────────────────────────────────────────────────

class TestHtmlExport:
    def test_is_complete_html_document(self):
        """완전한 HTML 문서 구조여야 한다."""
        content, mime, filename = export_record(SAMPLE_RECORD, "html")
        text = content.decode("utf-8")
        assert "<!DOCTYPE html>" in text
        assert "<html" in text
        assert "<meta charset" in text.lower()
        assert "</body>" in text
        assert mime == "text/html; charset=utf-8"
        assert filename.endswith(".html")

    def test_markdown_field_rendered_to_html_tags(self):
        """markdown hint 필드가 GFM 태그(<h1>, <strong> 등)로 렌더링된다."""
        content, _, _ = export_record(
            SAMPLE_RECORD, "html", viewer_hints=VIEWER_HINTS
        )
        text = content.decode("utf-8")
        # GFM # 안녕하세요 → <h1>
        assert "<h1>" in text or "<h1 " in text
        # **굵은 텍스트** → <strong>
        assert "<strong>" in text

    def test_field_html_single_card(self):
        """field 지정 시 해당 필드만 렌더링."""
        content, _, _ = export_record(
            SAMPLE_RECORD, "html", field="answer_md", viewer_hints=VIEWER_HINTS
        )
        text = content.decode("utf-8")
        assert "<h1>" in text or "<h1 " in text
        # 다른 필드 레이블은 없어야 함
        assert "<dt>score" not in text

    def test_no_field_shows_all_keys(self):
        """field 미지정 시 모든 키가 dl 에 포함된다."""
        content, _, _ = export_record(
            SAMPLE_RECORD, "html", viewer_hints=VIEWER_HINTS
        )
        text = content.decode("utf-8")
        assert "title" in text
        assert "answer_md" in text
        assert "score" in text


# ── xlsx 변환 ────────────────────────────────────────────────────────────────

class TestXlsxExport:
    def test_field_single_cell(self):
        """field 지정 시 헤더=field, 값=string."""
        pytest.importorskip("openpyxl")
        import openpyxl

        content, mime, filename = export_record(
            SAMPLE_RECORD, "xlsx", field="title"
        )
        assert mime.startswith("application/vnd.openxmlformats")
        assert filename.endswith(".xlsx")

        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "title"
        assert ws.cell(row=2, column=1).value == "샘플 제목"

    def test_no_field_all_columns(self):
        """field 미지정 시 첫 행 = 키, 두 번째 행 = 값."""
        pytest.importorskip("openpyxl")
        import openpyxl

        content, _, _ = export_record(SAMPLE_RECORD, "xlsx")
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active

        headers = [ws.cell(row=1, column=c).value for c in range(1, 4)]
        assert "title" in headers
        assert "answer_md" in headers
        assert "score" in headers

        # 값 행에 실제 값이 있는지 확인
        values_row = [ws.cell(row=2, column=c).value for c in range(1, 4)]
        assert any("샘플 제목" == v for v in values_row)
        assert any("42" == str(v) or v == 42 or v == "42" for v in values_row)

    def test_bytes_output(self):
        """반환값이 bytes 여야 한다."""
        pytest.importorskip("openpyxl")
        content, _, _ = export_record(SAMPLE_RECORD, "xlsx")
        assert isinstance(content, bytes)
        # xlsx magic bytes (PK zip)
        assert content[:2] == b"PK"


# ── 유효성 / 엣지케이스 ────────────────────────────────────────────────────

class TestValidation:
    def test_invalid_format_raises(self):
        """지원하지 않는 포맷은 ValueError 발생."""
        with pytest.raises(ValueError, match="지원하지 않는 포맷"):
            export_record(SAMPLE_RECORD, "pdf")

    def test_empty_data_no_crash(self):
        """data 가 빈 dict 여도 크래시 없이 반환."""
        rec = {**SAMPLE_RECORD, "data": {}}
        for fmt in ("md", "csv", "html"):
            content, _, _ = export_record(rec, fmt)
            assert isinstance(content, bytes)

    def test_no_viewer_hints_uses_plain(self):
        """viewer_hints 미지정 시 plain 직렬화."""
        content, _, _ = export_record(SAMPLE_RECORD, "md", viewer_hints=None)
        text = content.decode("utf-8")
        assert "title" in text
